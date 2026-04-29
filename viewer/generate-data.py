#!/usr/bin/env python3
"""Generate viewer/data.json from the James Waring spreadsheet."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "James Waring current version final.xlsx"
OUTPUT = Path(__file__).resolve().parent / "data.json"
THUMBNAILS = ROOT / "Waring-thumbnails586"
BIGGER_IMAGES = ROOT / "Waring-referenceFiles-Bigger"

FIELD_MAP = {
    "File Name": "fileName",
    "Thumbnail Link": "thumbnailLink",
    "Name / Caption": "nameCaption",
    "Photo": "photo",
    "Year": "year",
    "Copyright": "copyright",
    "Description": "description",
    "Comment": "comment",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def image_lookup(folder: Path) -> dict[str, str]:
    if not folder.exists():
        return {}

    return {path.name.casefold(): path.name for path in folder.iterdir() if path.is_file()}


def matching_image(file_name: str, lookup: dict[str, str]) -> str:
    return lookup.get(file_name.casefold(), "")


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)
    sheet = workbook.active
    headers = [clean(cell.value) for cell in sheet[1]]
    thumbnail_lookup = image_lookup(THUMBNAILS)
    bigger_lookup = image_lookup(BIGGER_IMAGES)
    rows = []

    for spreadsheet_row, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        item = {"spreadsheetRow": spreadsheet_row}

        for index, cell in enumerate(row):
            if index >= len(headers):
                continue

            key = FIELD_MAP.get(headers[index])
            if not key:
                continue

            value = clean(cell.value)
            if key == "thumbnailLink" and cell.hyperlink:
                value = cell.hyperlink.target
            item[key] = value

        if item.get("fileName"):
            file_name = item["fileName"]
            item["localThumbnail"] = matching_image(file_name, thumbnail_lookup)
            item["localImage"] = matching_image(file_name, bigger_lookup)
            rows.append(item)

    OUTPUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Wrote {len(rows)} rows to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
